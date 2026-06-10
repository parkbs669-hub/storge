# -*- coding: utf-8 -*-
"""
테니스 자동 포스팅 - 엑셀 본문 + 로컬 이미지 첨부 버전 (new_beom_sports)
- posting_naver.xlsx에서 제목/본문 읽기 (뉴스 수집/AI 생성 없음)
- 본문을 HTML로 변환 후 Tistory에 포스팅
- 로컬 이미지 폴더(테니스 블로그 이미지)에서 글 맥락에 맞는 이미지 선택
- 포스팅 엔진: Tistory Selenium (Kakao 로그인)
"""

import sys
import io
import re
import os
import json
import time
import random
import base64
import openpyxl
from pathlib import Path
from datetime import datetime

_BASE = Path(__file__).parent
_IMAGE_DIR = Path(r'C:\Users\박범서\claud\images')
_USED_IMAGES_FILE = _BASE / 'used_images_tistory.json'


def _load_used_tistory() -> set:
    try:
        return set(json.loads(_USED_IMAGES_FILE.read_text(encoding='utf-8')))
    except Exception:
        return set()


def _save_used_tistory(names: list):
    existing = _load_used_tistory()
    existing.update(names)
    out = list(existing)[-100:]
    _USED_IMAGES_FILE.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding='utf-8')


def _pick_female_closing_image() -> str | None:
    """선수_여성 이미지 중 이전에 쓰지 않은 것 1장 반환."""
    if not _IMAGE_DIR.exists():
        return None
    recently_used = _load_used_tistory()
    fresh = [
        f for f in _IMAGE_DIR.iterdir()
        if '선수_여성' in f.name and f.suffix.lower() in ('.jpg', '.jpeg', '.png')
        and f.name not in recently_used and f.stat().st_size > 50 * 1024
    ]
    if fresh:
        return str(random.choice(fresh))
    fallback = [
        f for f in _IMAGE_DIR.iterdir()
        if '선수_여성' in f.name and f.suffix.lower() in ('.jpg', '.jpeg', '.png')
        and f.stat().st_size > 50 * 1024
    ]
    return str(random.choice(fallback)) if fallback else None


def _img_to_html(img_path: str) -> str:
    """이미지 파일 → base64 인라인 HTML (항상 JPEG 700px/quality 65로 압축)"""
    try:
        from PIL import Image as _Img
        import io as _io
        with _Img.open(img_path) as im:
            if im.width > 700:
                ratio = 700 / im.width
                im = im.resize((700, int(im.height * ratio)), _Img.LANCZOS)
            buf = _io.BytesIO()
            im.convert('RGB').save(buf, format='JPEG', quality=65)
            b64 = base64.b64encode(buf.getvalue()).decode()
        src = f'data:image/jpeg;base64,{b64}'
    except Exception:
        src = f'data:image/jpeg;base64,{base64.b64encode(Path(img_path).read_bytes()).decode()}'
    return (f'<div style="margin:20px 0;border-radius:10px;overflow:hidden;'
            f'box-shadow:0 3px 12px rgba(0,0,0,.12);text-align:center;">'
            f'<img src="{src}" alt="테니스 선수" style="max-width:100%;height:auto;display:block;margin:0 auto;"></div>')


if sys.stdout.encoding and sys.stdout.encoding.lower() in ['cp949', 'euc-kr']:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.common.exceptions import StaleElementReferenceException
from webdriver_manager.chrome import ChromeDriverManager

TISTORY_BLOG = 'beomsports'
KAKAO_ID     = 'parkbs669@gmail.com'
KAKAO_PW     = 'Bs1809bs01!'

# ══════════════════════════════════════════════
#  ★ 설정값 ★
# ══════════════════════════════════════════════
EXCEL_FILE      = r"C:\Users\박범서\claud\네이버\posting_naver.xlsx"
USED_POSTS_FILE = r"C:\Users\박범서\claud\티스토리\used_posts_tistory.json"
CATEGORY_MAP = {
    "스트링":    "스트링 리뷰 & 팁",
    "텐션":      "스트링 리뷰 & 팁",
    "폴리":      "스트링 리뷰 & 팁",
    "나일론":    "스트링 리뷰 & 팁",
    "거트":      "스트링 리뷰 & 팁",
    "알루파워":  "스트링 리뷰 & 팁",
    "럭실론":    "스트링 리뷰 & 팁",
    "바볼랏":    "스트링 리뷰 & 팁",
    "솔린코":    "스트링 리뷰 & 팁",
    "스트링머신": "스트링 리뷰 & 팁",
    "대회":      "테니스 대회 & 뉴스",
    "토너먼트":  "테니스 대회 & 뉴스",
    "ATP":       "테니스 대회 & 뉴스",
    "WTA":       "테니스 대회 & 뉴스",
    "우승":      "테니스 대회 & 뉴스",
    "결승":      "테니스 대회 & 뉴스",
    "준결승":    "테니스 대회 & 뉴스",
    "마스터스":  "테니스 대회 & 뉴스",
    "그랜드슬램": "테니스 대회 & 뉴스",
    "롤랑가로스": "테니스 대회 & 뉴스",
    "윔블던":    "테니스 대회 & 뉴스",
    "US오픈":    "테니스 대회 & 뉴스",
    "호주오픈":  "테니스 대회 & 뉴스",
    "로마":      "테니스 대회 & 뉴스",
    "마드리드":  "테니스 대회 & 뉴스",
    "파리":      "테니스 대회 & 뉴스",
    "시너":      "테니스 대회 & 뉴스",
    "알카라스":  "테니스 대회 & 뉴스",
    "조코비치":  "테니스 대회 & 뉴스",
    "스비아텍":  "테니스 대회 & 뉴스",
    "사발렌카":  "테니스 대회 & 뉴스",
    "도서":      "테니스 도서 & 매거진",
    "매거진":    "테니스 도서 & 매거진",
    "리포트":    "테니스 리포트 & 트렌드",
    "트렌드":    "테니스 리포트 & 트렌드",
}
RECENT_LIMIT = 20
# ══════════════════════════════════════════════

# 디자인 테마
_C_PRIMARY  = "#2e7d32"
_C_DARK     = "#1b5e20"
_C_ACCENT   = "#43a047"
_C_LIGHT    = "#e8f5e9"
_C_INTERNAL = "#f1f8e9"

# ── 이미지 카테고리 매핑
IMAGE_CATEGORIES = {
    "스트링":  ["스트링", "스트링머신", "텐션"],
    "선수":    ["선수", "임팩트", "슬라이드"],
    "코트":    ["코트", "테니스_감성", "테니스_빈티지"],
    "장비":    ["장비", "라켓", "그립"],
}

KEYWORD_MAP = {
    "스트링":     ["스트링", "장비", "선수", "코트"],
    "텐션":       ["스트링", "장비", "선수", "코트"],
    "스트링머신":  ["스트링", "장비", "선수", "코트"],
    "폴리":       ["스트링", "장비", "선수", "코트"],
    "나일론":     ["스트링", "장비", "선수", "코트"],
    "거트":       ["스트링", "장비", "선수", "코트"],
    "알루파워":   ["스트링", "장비", "선수", "코트"],
    "럭실론":     ["스트링", "장비", "선수", "코트"],
    "바볼랏":     ["스트링", "장비", "선수", "코트"],
    "솔린코":     ["스트링", "장비", "선수", "코트"],
    "선수":       ["선수", "코트", "스트링", "장비"],
    "동호인":     ["선수", "코트", "스트링", "장비"],
    "코트":       ["코트", "선수", "스트링", "장비"],
    "클레이":     ["코트", "선수", "스트링", "장비"],
    "라켓":       ["장비", "스트링", "선수", "코트"],
    "장비":       ["장비", "스트링", "선수", "코트"],
    "그립":       ["장비", "스트링", "선수", "코트"],
}

# 영문 prefix → 한국어 하위폴더명 매핑 (images/ 아래 폴더)
_EN_PREFIX_TO_KO_FOLDER: dict[str, str] = {
    'Jannik_Sinner':        '야닉 시너',
    'Carlos_Alcaraz':       '카를로스 알카라즈',
    'Novak_Djokovic':       '조코비치',
    'Daniil_Medvedev':      '다닐 메드베데프',
    'Alexander_Zverev':     '알렉산더 즈베레프',
    'Andrey_Rublev':        '안드레이 루블레프',
    'Casper_Ruud':          '캐스퍼 루드',
    'Arthur_Fils':          '아르투르 피스',
    'Flavio_Cobolli':       '플라비오 코볼리',
    'Taylor_Fritz':         '테일러 프리츠',
    'Ben_Shelton':          '벤 쉘튼',
    'Lorenzo_Musetti':      '로렌조 무세티',
    'Alexander_Bublik':     '알렉산더 부블릭',
    'Alex_de':              '알렉스 드 미노',
    'Joao_Fonseca':         None,
    'Aryna_Sabalenka':      '아리나 사발렌카',
    'Iga_Swiatek':          '이가 시비옹테크',
    'Mirra_Andreeva':       '미라 안드레예바',
    'Elena_Rybakina':       '엘레나 리바키나',
    'Coco_Gauff':           '코코 고프',
    'Naomi_Osaka':          '나오미 오사카',
    'Marta_Kostyuk':        '마르타 코스튜크',
    'Elina_Svitolina':      '엘리나 스비톨리나',
    'Holger_Rune':          None,
    'Karen_Khachanov':      '카렌 하차노프',
    'Karolina_Muchova':     '카롤리나 무호바',
}

# 한국어 선수명 → 이미지 파일명 prefix 매핑
_PLAYER_FILEMAP = {
    "시너": "Jannik_Sinner", "야닉시너": "Jannik_Sinner", "야닉 시너": "Jannik_Sinner",
    "알카라스": "Carlos_Alcaraz", "알카라즈": "Carlos_Alcaraz",
    "조코비치": "Novak_Djokovic", "메드베데프": "Daniil_Medvedev",
    "즈베레프": "Alexander_Zverev", "지베레프": "Alexander_Zverev",
    "루블레프": "Andrey_Rublev", "루블료프": "Andrey_Rublev",
    "루드": "Casper_Ruud", "치치파스": "Stefanos_Tsitsipas",
    "루네": "Holger_Rune", "필스": "Arthur_Fils", "코볼리": "Flavio_Cobolli",
    "프리츠": "Taylor_Fritz", "셸튼": "Ben_Shelton", "셸턴": "Ben_Shelton", "드레이퍼": "Jack_Draper",
    "사발렌카": "Aryna_Sabalenka", "사바렌카": "Aryna_Sabalenka",
    "스비아텍": "Iga_Swiatek", "스비온텍": "Iga_Swiatek",
    "안드레에바": "Mirra_Andreeva", "안드레예바": "Mirra_Andreeva",
    "리바키나": "Elena_Rybakina", "고프": "Coco_Gauff",
    "칼린스카야": "Anna_Kalinskaya",
    "무세티": "Lorenzo_Musetti", "로렌초 무세티": "Lorenzo_Musetti", "로렌조 무세티": "Lorenzo_Musetti",
    "페더러": "Roger_Federer", "로저 페더러": "Roger_Federer",
    "후르카치": "Hubert_Hurkacz", "허버트 후르카치": "Hubert_Hurkacz",
    "폰세카": "Joao_Fonseca", "주앙 폰세카": "Joao_Fonseca",
    "샤포발로프": "Denis_Shapovalov",
    "세룬돌로": "Francisco_Cerundolo", "세룬돌": "Francisco_Cerundolo",
    "슈나이더": "Diana_Shnaider",
    "스비톨리나": "Elina_Svitolina",
    "코스튜크": "Marta_Kostyuk",
    "데 미나우르": "Alex_de", "알렉스 데 미나우르": "Alex_de",
    "구연우": "Yeonwoo_Gu",
    "플리스코바": "Karolina_Pliskova",
}


def _parse_postmeta(text: str) -> dict:
    """[POSTMETA]...[/POSTMETA] 블록에서 주선수/상대선수 파싱."""
    m = re.search(r'\[POSTMETA\](.*?)\[/POSTMETA\]', text, re.DOTALL | re.IGNORECASE)
    if not m:
        return {}
    result = {}
    for line in m.group(1).splitlines():
        if '주선수:' in line:
            result['main'] = line.split('주선수:')[1].strip()
        elif '상대선수:' in line:
            result['opponent'] = line.split('상대선수:')[1].strip()
    return result


def _resolve_postmeta_players(meta: dict) -> list:
    """[POSTMETA] 한국어명 → 영문 prefix 리스트. 주선수 먼저."""
    result = []
    for key in ('main', 'opponent'):
        name = meta.get(key, '').strip()
        if not name:
            continue
        en = _PLAYER_FILEMAP.get(name)
        if not en:
            for ko, prefix in _PLAYER_FILEMAP.items():
                if ko in name or name in ko:
                    en = prefix
                    break
        if en and en not in result:
            result.append(en)
    return result


def _pick_player_images(title: str, content: str, count: int, recently_used, forced_players: list = None) -> list:
    """본문 선수명 감지 → 선수 이미지 우선 선택.
    두 선수가 함께 등장: joint photo 먼저, 없으면 교대로. 단일 선수: 해당 선수 이미지만.
    """
    img_dir = _IMAGE_DIR
    if not img_dir.exists():
        return []

    if forced_players:
        # [POSTMETA]로 명시된 선수: recently_used 무시하고 모두 메인 취급
        players = forced_players[:]
        title_players: set = set(forced_players)
    else:
        # 빈도 기반 플레이어 감지: 제목 언급 3배 가중치, 본문 언급 1배
        player_scores: dict[str, int] = {}
        for ko, en in _PLAYER_FILEMAP.items():
            t_cnt = title.count(ko) * 3
            b_cnt = content.count(ko)
            score = t_cnt + b_cnt
            if score > 0:
                if en not in player_scores or score > player_scores[en]:
                    player_scores[en] = score
        players = sorted(player_scores, key=lambda e: -player_scores[e])
        if not players:
            return []

        # 제목에 직접 언급된 선수는 recently_used 필터 무시
        title_players = set()
        for ko, en in _PLAYER_FILEMAP.items():
            if ko in title:
                title_players.add(en)

    def _files_for(prefix: str, used_names: set) -> list:
        is_main = prefix in title_players
        all_matches: list = []
        # 루트 플랫 파일
        for f in img_dir.iterdir():
            if (f.is_file() and f.name.startswith(prefix)
                    and f.suffix.lower() in ('.jpg', '.jpeg', '.png')
                    and f.name not in used_names):
                all_matches.append(f)
        # 한글 하위폴더 파일
        ko_folder = _EN_PREFIX_TO_KO_FOLDER.get(prefix)
        if ko_folder:
            ko_dir = img_dir / ko_folder
            if ko_dir.is_dir():
                for f in ko_dir.iterdir():
                    if (f.is_file()
                            and f.suffix.lower() in ('.jpg', '.jpeg', '.png')
                            and f.name not in used_names):
                        all_matches.append(f)
        fresh = [f for f in all_matches if f.name not in recently_used]
        stale = [f for f in all_matches if f.name in recently_used]
        random.shuffle(fresh)
        random.shuffle(stale)
        if is_main:
            return [str(f) for f in fresh + stale]
        else:
            return [str(f) for f in fresh]

    selected: list = []
    used_names: set = set()

    if len(players) >= 2:
        # 두 선수가 함께 등장 → joint photo 먼저 시도
        p1, p2 = players[0], players[1]
        p1_is_main = p1 in title_players
        p2_is_main = p2 in title_players
        joint = [str(f) for f in img_dir.iterdir()
                 if f.suffix.lower() in ('.jpg', '.jpeg', '.png')
                 and p1 in f.name and p2 in f.name
                 and ((p1_is_main or p2_is_main) or f.name not in recently_used)]
        if joint:
            selected.append(joint[0])
            used_names.add(os.path.basename(joint[0]))
        # joint photo 없거나 추가 슬롯 필요 → 두 선수 교대 선택
        p1_imgs = _files_for(p1, used_names)
        p2_imgs = _files_for(p2, used_names)
        i1 = i2 = 0
        while len(selected) < count and (i1 < len(p1_imgs) or i2 < len(p2_imgs)):
            if i1 < len(p1_imgs):
                img = p1_imgs[i1]; i1 += 1
                if os.path.basename(img) not in used_names:
                    selected.append(img); used_names.add(os.path.basename(img))
            if len(selected) < count and i2 < len(p2_imgs):
                img = p2_imgs[i2]; i2 += 1
                if os.path.basename(img) not in used_names:
                    selected.append(img); used_names.add(os.path.basename(img))
    else:
        # 단일 주요 선수
        for img in _files_for(players[0], used_names):
            selected.append(img); used_names.add(os.path.basename(img))
            if len(selected) >= count:
                break

    # 3번째 이후 선수 이미지로 나머지 채우기
    for player in players[2:]:
        if len(selected) >= count:
            break
        for img in _files_for(player, used_names):
            selected.append(img); used_names.add(os.path.basename(img))
            break

    return selected[:count]


COMMON_FOOTER_HTML = f"""
<div style="margin-top:40px;padding:25px;background:{_C_INTERNAL};
            border-radius:12px;border:1px solid #c8e6c9;box-shadow:0 2px 8px rgba(0,0,0,0.05);">
  <h3 style="margin:0 0 15px;color:{_C_DARK};font-size:18px;">🔍 Beom Sports 인기 콘텐츠</h3>
  <ul style="list-style:none;padding:0;margin:0;line-height:2.0;font-size:15px;">
    <li style="margin-bottom:8px;">🔗 <a href="https://beomsports.tistory.com/category/%ED%85%8C%EB%8B%88%EC%8A%A4%20%EB%8C%80%ED%9A%8C%20%26%20%EB%89%B4%EC%8A%A4" target="_blank" style="color:{_C_PRIMARY};text-decoration:none;font-weight:bold;">테니스 최신뉴스</a></li>
    <li style="margin-bottom:8px;">🔗 <a href="https://beomsports.tistory.com/category/%EC%8A%A4%ED%8A%B8%EB%A7%81%20%EB%A6%AC%EB%B7%B0%20%26%20%ED%8C%81" target="_blank" style="color:{_C_PRIMARY};text-decoration:none;font-weight:bold;">테니스 스트링 선택</a></li>
    <li>🔗 <a href="https://beomsports.tistory.com/category/%ED%85%8C%EB%8B%88%EC%8A%A4%20%EB%8F%84%EC%84%9C%20%26%20%EB%A7%A4%EA%B1%B0%EC%A7%84" target="_blank" style="color:{_C_PRIMARY};text-decoration:none;font-weight:bold;">테니스 도서 추천 리뷰</a></li>
  </ul>
</div>
<div style="margin-top:24px;padding:24px 20px;background:linear-gradient(135deg,{_C_DARK} 0%,{_C_ACCENT} 100%);
            border-radius:14px;text-align:center;box-shadow:0 4px 16px rgba(0,0,0,0.12);">
  <p style="font-size:16px;color:#fff;font-weight:800;margin:0 0 6px;">🎾 테니스가 좋다면? 범스포츠를 구독하세요!</p>
  <p style="font-size:13px;color:#c8e6c9;margin:0 0 10px;">매일 업데이트되는 글로벌 테니스 최신 소식.</p>
  <p style="font-size:12px;color:#a5d6a7;margin:0;">👉 좋아요와 댓글이 큰 힘이 됩니다 😊</p>
</div>"""


def get_driver():
    options = Options()
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.add_experimental_option('excludeSwitches', ['enable-automation'])
    options.add_experimental_option('useAutomationExtension', False)
    options.add_argument('--window-size=1400,900')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-gpu')
    
    # ── [추가] 크롬 세션 쿠키 저장을 위한 독립 프로필 경로 지정 ──
    profile_path = r"C:\Users\박범서\claud\티스토리\chrome_profile"
    os.makedirs(profile_path, exist_ok=True)
    options.add_argument(f"--user-data-dir={profile_path}")
    
    service = ChromeService(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    return driver


def _kakao_login(driver):
    """카카오 계정으로 티스토리 로그인 (세션 재사용 지원 및 2단계 대기)"""
    wait = WebDriverWait(driver, 35)
    
    # 1. 먼저 메인 페이지로 이동하여 기존 세션이 유지되고 있는지 체크
    driver.get('https://www.tistory.com/')
    time.sleep(3)
    
    # 로그인 상태 체크 (현재 URL이 tistory 도메인이고 login/auth 페이지가 아닌 경우)
    if "tistory.com" in driver.current_url and "login" not in driver.current_url.lower() and "auth" not in driver.current_url.lower():
        try:
            # 비로그인 시 나타나는 로그인 유도 버튼 여부 확인
            driver.find_element(By.CSS_SELECTOR, "a.btn_login, a[href*='auth/login']")
        except Exception:
            print('   ✅ [세션 재사용] 기존 크롬 프로필에 저장된 로그인 세션이 유효합니다. 카카오 로그인을 건너뜁니다.')
            return

    # 2. 비로그인 상태일 경우 로그인 절차 수행
    driver.get('https://www.tistory.com/auth/login')
    time.sleep(4)
    kakao_btn = None
    for sel in ['a.link_kakao_id', "a[href*='accounts.kakao.com']",
                'a.btn_login.kakao_account', "a[data-social-login='kakao']"]:
        try:
            kakao_btn = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, sel)))
        except Exception:
            continue
        else:
            break
    if not kakao_btn:
        kakao_btn = wait.until(EC.element_to_be_clickable((
            By.XPATH, "//a[.//span[contains(text(),'카카오')] or contains(@class,'kakao')]"
        )))
    driver.execute_script('arguments[0].click();', kakao_btn)
    time.sleep(4)
    
    # ── [추가] 간편 로그인(카카오계정 선택) 처리 ──
    account_selected = False
    try:
        # 이메일 텍스트가 화면에 노출되는지 확인 (최대 10초 대기)
        WebDriverWait(driver, 10).until(
            lambda d: "로그인할 카카오계정 선택" in d.page_source or KAKAO_ID in d.page_source
        )
        print("   🎯 간편 로그인(카카오계정 선택) 화면이 감지되었습니다.")
        
        # KAKAO_ID를 포함하는 element 클릭
        target = None
        for xpath in [
            f"//button[.//span[contains(text(), '{KAKAO_ID}')]]",
            f"//span[contains(text(), '{KAKAO_ID}')]",
            f"//*[contains(text(), '{KAKAO_ID}')]"
        ]:
            try:
                target = driver.find_element(By.XPATH, xpath)
                if target:
                    break
            except Exception:
                continue
        if target:
            driver.execute_script("arguments[0].click();", target)
            print("   ✅ 간편 로그인 계정 선택 및 클릭 완료")
            account_selected = True
            time.sleep(3)
        else:
            print("   ⚠️ 간편 로그인 대상 계정 요소를 찾을 수 없어 일반 로그인으로 진행합니다.")
    except Exception:
        # 간편 로그인 화면이 안 떴거나 에러가 나면 일반 로그인으로 패스
        print("   ℹ️ 일반 카카오 로그인 폼으로 진행합니다.")
        
    if not account_selected:
        id_input = wait.until(EC.presence_of_element_located((
            By.CSS_SELECTOR, "input#loginId, input[name='loginId'], input[type='email']"
        )))
        id_input.clear()
        id_input.send_keys(KAKAO_ID)
        time.sleep(0.5)
        pw_input = wait.until(EC.presence_of_element_located((
            By.CSS_SELECTOR, "input#password, input[name='password'], input[type='password']"
        )))
        pw_input.clear()
        pw_input.send_keys(KAKAO_PW)
        time.sleep(0.5)
        
        # ── [추가] 로그인 상태 유지 체크박스 클릭 ──
        try:
            keep_login = None
            for sel in ["input[name='saveSignedIn']", "input[id^='saveSignedIn']", "input#keepLogin", "input[name='keepLogin']", "input#saveId"]:
                try:
                    keep_login = driver.find_element(By.CSS_SELECTOR, sel)
                    if keep_login:
                        break
                except Exception:
                    continue
            if keep_login:
                if not keep_login.is_selected():
                    driver.execute_script('arguments[0].click();', keep_login)
                    print("   ✅ 로그인 상태 유지 체크박스 선택 완료")
            else:
                print("   ⚠️ 로그인 상태 유지 체크박스를 찾을 수 없습니다.")
        except Exception as e:
            print(f"   ⚠️ 로그인 상태 유지 선택 시도 중 오류: {e}")
            
        time.sleep(0.5)
        login_btn = wait.until(EC.element_to_be_clickable((
            By.CSS_SELECTOR, "button[type='submit'].btn_g, button[type='submit']"
        )))
        driver.execute_script('arguments[0].click();', login_btn)
        time.sleep(5)
    
    # ── [추가] 2단계 핸드폰 인증 대기 로직 ──
    print("   🔑 로그인 버튼 클릭 완료. 2단계 핸드폰 보안 인증 단계를 대기합니다.")
    try:
        print("   👉 핸드폰에 전송된 카카오 인증 승인 요청을 수락하거나 번호를 입력해 주세요. (제한시간 120초)")
        # 로그인 성공 후 tistory 메인 도메인(login/auth 단어가 빠진 URL)으로 복귀할 때까지 최대 120초 대기
        WebDriverWait(driver, 120).until(
            lambda d: "tistory.com" in d.current_url and "login" not in d.current_url.lower() and "auth" not in d.current_url.lower()
        )
        print('   ✅ 카카오 로그인 및 2단계 인증 승인 완료')
    except Exception as e:
        print("   ❌ 2단계 인증 시간 초과 또는 로그인 실패. 스크립트를 수동으로 재실행해주세요.")
        raise e


# ── 이미지 파일 수집
def get_all_images():
    categorized = {cat: [] for cat in IMAGE_CATEGORIES}
    uncategorized = []
    if not _IMAGE_DIR.exists():
        print(f"⚠️  이미지 폴더 없음: {_IMAGE_DIR}")
        return categorized, uncategorized
    for fname in os.listdir(str(_IMAGE_DIR)):
        if not fname.lower().endswith((".jpg", ".jpeg", ".png")):
            continue
        full_path = str(_IMAGE_DIR / fname)
        matched = False
        for cat, keywords in IMAGE_CATEGORIES.items():
            if any(kw in fname for kw in keywords):
                categorized[cat].append(full_path)
                matched = True
                break
        if not matched:
            uncategorized.append(full_path)
    return categorized, uncategorized


def select_images(title: str, content: str = "", count: int = 5, forced_players: list = None) -> list:
    """제목/본문 키워드에 맞는 로컬 이미지 count장 선택."""
    categorized, uncategorized = get_all_images()
    combined_text = title + " " + content[:300]
    selected = []
    recently_used = _load_used_tistory()

    def prefer_fresh(imgs):
        fresh = [i for i in imgs if os.path.basename(i) not in recently_used]
        stale = [i for i in imgs if os.path.basename(i) in recently_used]
        random.shuffle(fresh)
        random.shuffle(stale)
        return fresh + stale

    # ── 선수 이미지 우선 선택 (선수명이 감지되면 해당 선수 사진으로 채움)
    player_imgs = _pick_player_images(title, content, count, recently_used, forced_players=forced_players)
    if player_imgs:
        selected = player_imgs[:]
        if len(selected) < count:
            all_flat = prefer_fresh([i for imgs in categorized.values() for i in imgs] + uncategorized)
            for img in all_flat:
                if img not in selected:
                    selected.append(img)
                if len(selected) >= count:
                    break
        print(f"🖼️  이미지 {len(selected)}장 선택:")
        for p in selected:
            print(f"   - {os.path.basename(p)}")
        return selected[:count]

    string_article_keywords = ["스트링", "텐션", "폴리", "나일론", "거트", "알루파워",
                                "럭실론", "바볼랏", "솔린코", "요넥스", "스트링머신", "헤드", "윌슨"]
    is_string_article = any(kw in title for kw in string_article_keywords)

    if is_string_article:
        string_imgs = prefer_fresh(categorized.get("스트링", []))
        selected = string_imgs[:count]
        if len(selected) < count:
            extra = prefer_fresh([i for i in categorized.get("장비", []) if i not in selected])
            selected += extra[:count - len(selected)]
        if len(selected) < count:
            all_imgs = prefer_fresh([p for imgs in categorized.values() for p in imgs] + uncategorized)
            for img in all_imgs:
                if img not in selected:
                    selected.append(img)
                if len(selected) >= count:
                    break
        print(f"🎯 스트링 글 → 스트링 카테고리 이미지만 선택")
    else:
        keyword_matched = False
        brand_keywords = ["알루파워", "럭실론", "바볼랏", "솔린코", "요넥스", "헤드", "윌슨",
                          "시그넘", "테크니화이버", "폴리투어", "rpm", "hyperg", "alu"]
        matched_brands = [bkw for bkw in brand_keywords if bkw in title.lower()]
        if matched_brands:
            keyword_matched = True
            per_brand = max(1, 3 // len(matched_brands))
            for bkw in matched_brands:
                brand_imgs = []
                for imgs in categorized.values():
                    for img in imgs:
                        if bkw in os.path.basename(img).lower() and img not in selected:
                            brand_imgs.append(img)
                brand_imgs = prefer_fresh(brand_imgs)
                selected.extend(brand_imgs[:per_brand])

        priority = []
        for kw, cats in KEYWORD_MAP.items():
            if kw in combined_text:
                for c in cats:
                    if c not in priority:
                        priority.append(c)
        for cat in IMAGE_CATEGORIES:
            if cat not in priority:
                priority.append(cat)
        for cat in priority:
            if len(selected) >= count:
                break
            candidates = prefer_fresh([i for i in categorized.get(cat, []) if i not in selected])
            if candidates:
                selected.append(candidates[0])
                keyword_matched = True

        if not keyword_matched:
            default_img = _pick_female_closing_image()
            if default_img and default_img not in selected:
                selected.insert(0, default_img)
                print(f"   ℹ️  키워드 매칭 없음 → 기본 대표 이미지: {os.path.basename(default_img)}")

        all_imgs = [p for imgs in categorized.values() for p in imgs] + uncategorized
        all_imgs = prefer_fresh([i for i in all_imgs if i not in selected])
        for img in all_imgs:
            selected.append(img)
            if len(selected) >= count:
                break

    print(f"🖼️  이미지 {len(selected)}장 선택:")
    for p in selected:
        print(f"   - {os.path.basename(p)}")
    return selected


# ── STEP 1. 엑셀에서 글 데이터 읽기
def load_used_posts() -> list:
    if os.path.exists(USED_POSTS_FILE):
        try:
            with open(USED_POSTS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return []


def save_used_post(row_num: int):
    used = load_used_posts()
    if row_num not in used:
        used.append(row_num)
    with open(USED_POSTS_FILE, "w", encoding="utf-8") as f:
        json.dump(used, f, ensure_ascii=False, indent=2)


def read_post_from_excel():
    """posting_naver.xlsx에서 미사용 행을 랜덤 선택해 (row_num, title, body_html) 반환"""
    print(f"📊 엑셀 읽기: {EXCEL_FILE}")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    rows = []
    for row in range(2, ws.max_row + 1):
        title   = ws[f"A{row}"].value
        content = ws[f"B{row}"].value
        if title and content:
            rows.append((row, str(title).strip(), str(content).strip()))

    if not rows:
        print("❌ 포스팅할 데이터가 없습니다.")
        return None, None, None

    used_rows = load_used_posts()
    fresh_rows = [r for r in rows if r[0] not in used_rows]
    if not fresh_rows:
        print("🔄 모든 글 사용 완료 → 순환 초기화")
        with open(USED_POSTS_FILE, "w", encoding="utf-8") as f:
            json.dump([], f)
        fresh_rows = rows

    row_num, title, content = random.choice(fresh_rows)
    print(f"🎲 선택: {row_num}행 - {title[:40]}... (미사용 {len(fresh_rows)}개 중)")

    body_html = content_to_html(content)
    return row_num, title, body_html


def content_to_html(text: str) -> str:
    """엑셀 본문 텍스트를 가독성 좋은 HTML로 변환. 이미 HTML 태그가 있으면 그대로 반환."""
    if re.search(r"<(h[1-6]|p|div|ul|table)\b", text, re.IGNORECASE):
        return text

    CSS = """<style>
.post-wrap{font-family:'Apple SD Gothic Neo','Malgun Gothic',sans-serif;font-size:16px;
           line-height:1.95;color:#2d2d2d;max-width:740px;margin:0 auto;}
.post-wrap h2{font-size:20px;font-weight:700;color:#fff;margin:44px 0 16px;
              padding:14px 20px;background:linear-gradient(135deg,#1a6da3 0%,#2e86c1 100%);
              border-radius:8px;box-shadow:0 2px 8px rgba(46,134,193,0.25);}
.post-wrap h3{font-size:17px;font-weight:700;color:#1a6da3;margin:28px 0 10px;
              padding:8px 14px 8px 16px;border-left:4px solid #2e86c1;
              background:#f0f7fd;border-radius:0 6px 6px 0;}
.post-wrap p{font-size:16px;color:#333;margin:12px 0;line-height:1.95;}
.post-wrap .lead{font-size:16px;color:#444;line-height:2.0;padding:18px 22px;
                 background:#f8fbff;border:1px solid #d6eaf8;border-radius:8px;
                 margin:0 0 24px;}
.post-wrap .toc{background:#f8fbff;padding:18px 24px;border:1px solid #aed6f1;
                border-radius:8px;margin:20px 0 30px;}
.post-wrap .toc-title{font-weight:700;color:#1a6da3;font-size:15px;margin-bottom:10px;}
.post-wrap .toc ol{margin:0 0 0 18px;padding:0;}
.post-wrap .toc li{margin:5px 0;font-size:15px;}
.post-wrap .toc a{color:#2e86c1;text-decoration:none;}
.post-wrap .toc a:hover{text-decoration:underline;}
.post-wrap .lbl-tip{background:#e9f7ef;border-left:4px solid #27ae60;
                    padding:11px 16px;border-radius:0 6px 6px 0;margin:10px 0;font-size:15px;}
.post-wrap .lbl-warn{background:#fef9e7;border-left:4px solid #e67e22;
                     padding:11px 16px;border-radius:0 6px 6px 0;margin:10px 0;font-size:15px;}
.post-wrap .lbl-bad{background:#fdedec;border-left:4px solid #e74c3c;
                    padding:11px 16px;border-radius:0 6px 6px 0;margin:10px 0;font-size:15px;}
.post-wrap .lbl-strategy{background:#eaf2fb;border-left:4px solid #2e86c1;
                          padding:11px 16px;border-radius:0 6px 6px 0;margin:10px 0;font-size:15px;}
.post-wrap .lbl-reason{background:#f4f6f7;border-left:4px solid #7f8c8d;
                       padding:11px 16px;border-radius:0 6px 6px 0;margin:10px 0;font-size:15px;}
.post-wrap .lbl-def{background:#f0f3fa;border-left:4px solid #5d6d7e;
                    padding:11px 16px;border-radius:0 6px 6px 0;margin:8px 0;font-size:15px;}
.post-wrap .summary-box{background:#e8f8f5;border:1px solid #a9dfbf;
                         border-radius:8px;padding:20px 24px;margin:30px 0;}
.post-wrap .summary-box .sb-title{font-weight:700;color:#1e8449;font-size:17px;
                                   margin-bottom:10px;}
.post-wrap .key-point{background:#dbeafe;border-left:5px solid #1d4ed8;
                       padding:14px 18px;border-radius:0 8px 8px 0;margin:16px 0;
                       font-size:15px;color:#1e3a8a;font-weight:600;}
.post-wrap .highlight-box{background:#fef3c7;border-left:5px solid #d97706;
                            padding:12px 16px;border-radius:0 6px 6px 0;margin:10px 0;
                            font-size:15px;color:#78350f;}
.post-wrap .quote-box{background:#f3f4f6;border-left:4px solid #9ca3af;
                       padding:14px 20px;border-radius:0 6px 6px 0;margin:16px 0;
                       font-size:15px;color:#374151;font-style:italic;}
.post-wrap .grid-2col{display:grid;grid-template-columns:1fr 1fr;gap:12px;margin:16px 0;}
.post-wrap .grid-card{background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;
                       padding:14px 16px;font-size:14px;line-height:1.7;}
.post-wrap .grid-card .card-title{color:#1a6da3;font-weight:700;font-size:15px;
                                    display:block;margin-bottom:6px;}
.post-wrap .next-preview{background:#dcfce7;border:2px solid #86efac;border-radius:10px;
                           padding:18px 22px;margin:28px 0;}
.post-wrap .next-preview .np-title{font-weight:700;color:#15803d;font-size:16px;margin-bottom:8px;}
.post-wrap .source-box{background:#f8fafc;border:1px solid #e2e8f0;border-radius:6px;
                        padding:8px 14px;margin:16px 0;font-size:11px;color:#94a3b8;line-height:1.6;}
.post-wrap .source-box .src-title{font-weight:600;color:#94a3b8;font-size:11px;margin-bottom:4px;}
.post-wrap strong{color:#154360;}
table{width:100%;border-collapse:collapse;margin:20px 0;font-size:14px;}
table thead{background:#2e86c1;color:#fff;}
table th,table td{padding:11px 14px;text-align:left;border:1px solid #d6eaf8;}
table tbody tr:nth-child(even){background:#f0f7fd;}
</style>"""

    CIRCLE = "①②③④⑤⑥⑦⑧⑨⑩"

    LABEL_MAP = {
        "특징":      ("lbl-tip",      "✅ 특징"),
        "장점":      ("lbl-tip",      "👍 장점"),
        "이점":      ("lbl-tip",      "👍 이점"),
        "추천":      ("lbl-tip",      "⭐ 추천"),
        "주의":      ("lbl-warn",     "⚠️ 주의"),
        "주의점":    ("lbl-warn",     "⚠️ 주의"),
        "단점":      ("lbl-bad",      "❌ 단점"),
        "한계":      ("lbl-bad",      "❌ 한계"),
        "세팅 전략": ("lbl-strategy", "🎯 세팅 전략"),
        "세팅전략":  ("lbl-strategy", "🎯 세팅 전략"),
        "이유":      ("lbl-reason",   "💡 이유"),
    }

    def bold(s):
        return re.sub(r'\*\*(.+?)\*\*', r'<strong>\1</strong>', s)

    lines = text.strip().split("\n")

    first_idx = 0
    for idx, l in enumerate(lines):
        if l.strip():
            first_idx = idx
            break
    first = lines[first_idx].strip()
    is_title_line = bool(re.match(r'^\[.+\].+', first))
    start_idx = first_idx + 1 if is_title_line else first_idx

    parts = []
    toc_items = []
    sec_num = 0

    collect_mode = None
    collect_items = []
    collect_header = ""

    def flush_collect():
        nonlocal collect_mode, collect_items, collect_header
        if not collect_mode or not collect_items:
            collect_mode = None
            collect_items = []
            collect_header = ""
            return
        if collect_mode == 'grid':
            cards = ""
            for item in collect_items:
                sep = re.match(r'^([^:：]+)[：:]\s*(.+)$', item)
                if sep:
                    cards += (
                        f'<div class="grid-card">'
                        f'<span class="card-title">{sep.group(1).strip()}</span>'
                        f'{bold(sep.group(2).strip())}</div>'
                    )
                else:
                    cards += f'<div class="grid-card">{bold(item)}</div>'
            parts.append(
                f'<p style="font-weight:700;color:#1a6da3;margin:20px 0 8px;">{collect_header}</p>'
                f'<div class="grid-2col">{cards}</div>'
            )
        elif collect_mode == 'source':
            src_items = "".join(
                f'<p style="margin:3px 0;">• {bold(i)}</p>' for i in collect_items
            )
            parts.append(
                f'<div class="source-box">'
                f'<div class="src-title">📚 {collect_header}</div>'
                f'{src_items}</div>'
            )
        collect_mode = None
        collect_items = []
        collect_header = ""

    for raw in lines[start_idx:]:
        line = raw.strip()

        if not line:
            if collect_mode:
                flush_collect()
            continue

        # ## N. 제목 또는 ## 제목 형식 처리 (Claude 생성 글)
        mm = re.match(r'^##\s+(?:(\d+)\.\s+)?(.+)$', line)
        if mm:
            if collect_mode:
                collect_items.append(line)
                continue
            sec_num += 1
            sid = f"sec{sec_num}"
            raw_heading = mm.group(2)
            heading = bold(raw_heading)
            if not re.search(r'출처|참고|Reference', raw_heading, re.IGNORECASE):
                toc_items.append((sid, raw_heading))
            prefix = f"{mm.group(1)}. " if mm.group(1) else ""
            parts.append(f'<h2 id="{sid}">{prefix}{heading}</h2>')
            continue

        m = re.match(r'^(\d+)\.\s+(.+)$', line)
        if m:
            if collect_mode:
                collect_items.append(line)  # 수집 모드 중이면 항목으로 추가 (h2 변환 금지)
                continue
            sec_num += 1
            sid = f"sec{sec_num}"
            heading = bold(m.group(2))
            if not re.search(r'출처|참고|Reference', m.group(2), re.IGNORECASE):
                toc_items.append((sid, m.group(2)))
            parts.append(f'<h2 id="{sid}">{m.group(1)}. {heading}</h2>')
            continue

        if line and line[0] in CIRCLE:
            if collect_mode:
                collect_items.append(line)  # 수집 모드 중이면 항목으로 추가
                continue
            rest = line[1:].strip()
            if ":" in rest:
                t, s = rest.split(":", 1)
                parts.append(
                    f'<h3>{line[0]} {bold(t.strip())}'
                    f'<span style="color:#7f8c8d;font-weight:400;">: {bold(s.strip())}</span></h3>'
                )
            else:
                parts.append(f'<h3>{line[0]} {bold(rest)}</h3>')
            continue

        if collect_mode:
            collect_items.append(line)
            continue

        sum_match = re.match(r'^(결론|정리|요약)\s*[:：]\s*(.*)$', line)
        if sum_match:
            parts.append(
                f'<div class="summary-box">'
                f'<div class="sb-title">🏆 {sum_match.group(1)}</div>'
                f'<p style="margin:0">{bold(sum_match.group(2))}</p></div>'
            )
            continue

        m = re.match(r'^핵심\s*(문장|포인트|요소|내용)?\s*[:：]\s*(.+)$', line)
        if m:
            parts.append(f'<div class="key-point">💡 {bold(m.group(2))}</div>')
            continue

        m = re.match(r'^(중요|수치|데이터|통계|포인트)\s*[:：]\s*(.+)$', line)
        if m:
            icons = {"중요": "⭐", "수치": "📊", "데이터": "📈", "통계": "📉", "포인트": "📌"}
            icon = icons.get(m.group(1), "📌")
            parts.append(
                f'<div class="highlight-box">{icon} <strong>{m.group(1)}</strong>: {bold(m.group(2))}</div>'
            )
            continue

        m = re.match(r'^[>＞]\s*(.+)$', line)
        if m:
            parts.append(f'<div class="quote-box">❝ {bold(m.group(1))}</div>')
            continue
        m = re.match(r'^인용\s*[:：]\s*(.+)$', line)
        if m:
            parts.append(f'<div class="quote-box">❝ {bold(m.group(1))}</div>')
            continue

        m = re.match(r'^다음\s*(예고|글|포스팅|회차)?\s*[:：]\s*(.+)$', line)
        if m:
            label = m.group(1) or "예고"
            parts.append(
                f'<div class="next-preview">'
                f'<div class="np-title">🔜 다음 {label}</div>'
                f'<p style="margin:0">{bold(m.group(2))}</p></div>'
            )
            continue

        m = re.match(r'^국가별\s*(시각|현황|반응|분석|뷰)?\s*[:：]\s*(.*)$', line)
        if m:
            collect_mode = 'grid'
            collect_header = "🌍 국가별 " + (m.group(1) or "시각")
            if m.group(2):
                collect_items.append(m.group(2))
            continue

        m = re.match(r'^(출처|참고|Reference)\s*[:：]\s*(.*)$', line, re.IGNORECASE)
        if m:
            collect_mode = 'source'
            collect_header = m.group(1)
            if m.group(2):
                collect_items.append(m.group(2))
            continue

        matched = False
        for kw, (cls, icon) in LABEL_MAP.items():
            m = re.match(rf'^{re.escape(kw)}\s*[:：]\s*(.+)$', line)
            if m:
                parts.append(
                    f'<div class="{cls}"><strong>{icon}</strong>: {bold(m.group(1))}</div>'
                )
                matched = True
                break
        if matched:
            continue

        m = re.match(r'^([가-힣A-Za-z\s]{2,30}(?:\([^)]+\))?(?:\s*\d+%)?)\s*:\s*(.{15,})$', line)
        if m:
            term = m.group(1).strip()
            desc = bold(m.group(2).strip())
            parts.append(
                f'<div class="lbl-def"><strong>🔷 {term}</strong>: {desc}</div>'
            )
            continue

        parts.append(f'<p>{bold(line)}</p>')

    if collect_mode:
        flush_collect()

    toc_html = ""
    if len(toc_items) >= 2:
        items_html = "\n".join(
            f'<li><a href="#{sid}">{txt}</a></li>'
            for sid, txt in toc_items
        )
        toc_html = (
            f'<div class="toc">'
            f'<div class="toc-title">📋 목차</div>'
            f'<ol>{items_html}</ol>'
            f'</div>'
        )

    lead_html = f'<div class="lead">{bold(first)}</div>' if is_title_line and first else ""

    body = "\n".join(parts)
    return f'{CSS}\n<div class="post-wrap">\n{lead_html}\n{toc_html}\n{body}\n</div>'


def md_to_html(text: str) -> str:
    """마크다운 스타일 txt → 심플 HTML (## 헤더, --- 구분선, **볼드**, ①② 소제목 변환)"""
    CIRCLE = "①②③④⑤⑥⑦⑧⑨⑩"

    # 소프트 줄바꿈 합치기: 같은 문단 내에서 wrapping된 줄을 하나로 이어붙임
    _merged = []
    _buf: list = []
    for _l in text.splitlines():
        _s = _l.strip()
        _is_special = (
            not _s or _s.startswith('##') or _s.startswith('###') or
            _s == '---' or (_s and _s[0] in CIRCLE) or
            (_s and _s[0] in ('“', '”', '‘')) or
            bool(re.match(r'^\d+\. ', _s))
        )
        if _is_special:
            if _buf:
                _merged.append(' '.join(_buf))
                _buf = []
            _merged.append(_l)
        else:
            _buf.append(_s)
    if _buf:
        _merged.append(' '.join(_buf))
    text = '\n'.join(_merged)

    result = []
    h2_counter = 0
    for line in text.splitlines():
        stripped = line.strip()
        # 전부 # 기호인 장식 줄 제거
        if stripped and all(tok.startswith('#') for tok in stripped.split()):
            continue
        # --- 구분선
        if stripped == '---':
            result.append('<hr style="border:none;border-top:2px dashed #e2e8f0;margin:32px 0;">')
            continue
        # ### 소제목
        if line.startswith('### '):
            result.append(
                f'<h3 style="font-size:18px;font-weight:700;color:#1a3a6b;margin:24px 0 10px;'
                f'padding-left:12px;border-left:3px solid #c85000;">{line[4:]}</h3>'
            )
            continue
        # 숫자 형식 섹션 (1. 제목, 2. 제목, ...) → ## 와 동일하게 h2 처리
        _num_m = re.match(r'^(\d+)\. (.+)', stripped)
        if _num_m:
            h2_counter += 1
            heading_text = stripped
            result.append(
                f'<h2 id="sec{h2_counter}" style="font-size:21px;font-weight:800;color:#fff;'
                f'margin:44px 0 18px;padding:16px 22px;'
                f'background:linear-gradient(135deg,#1b3f7a 0%,#2e6da4 100%);'
                f'border-radius:8px;box-shadow:0 3px 12px rgba(27,63,122,0.25);">{heading_text}</h2>'
            )
            continue
        # ## 대제목 → 진한 배경 헤더
        if line.startswith('## '):
            h2_counter += 1
            heading_text = line[3:]
            result.append(
                f'<h2 id="sec{h2_counter}" style="font-size:21px;font-weight:800;color:#fff;'
                f'margin:44px 0 18px;padding:16px 22px;'
                f'background:linear-gradient(135deg,#1b3f7a 0%,#2e6da4 100%);'
                f'border-radius:8px;box-shadow:0 3px 12px rgba(27,63,122,0.25);">{heading_text}</h2>'
            )
            continue
        # 빈 줄
        if not stripped:
            continue
        # ①②③ 원문자 소제목
        if stripped[0] in CIRCLE:
            result.append(
                f'<h3 style="font-size:16px;font-weight:700;color:#1b3f7a;'
                f'margin:22px 0 10px;padding:10px 16px;'
                f'background:#eef4ff;border-left:4px solid #2e6da4;'
                f'border-radius:0 6px 6px 0;">{stripped}</h3>'
            )
            continue
        # 인용구: " 또는 " 로 시작하는 줄
        if stripped[0] in ('“', '”', '‘'):
            result.append(
                f'<div style="margin:16px 0 20px;padding:14px 20px;'
                f'background:#f8faff;border-left:4px solid #90a4ae;'
                f'border-radius:0 8px 8px 0;font-style:italic;'
                f'color:#455a64;font-size:15px;line-height:1.9;">{stripped}</div>'
            )
            continue
        # 본문: [출처] 작은 텍스트 + **볼드** 처리
        line_html = re.sub(r'\*\*(.+?)\*\*',
                           r'<strong style="color:#1b3f7a;font-weight:700;">\1</strong>', stripped)
        line_html = re.sub(r'\[([^\]]+)\]',
                           r'<sup style="color:#9e9e9e;font-size:11px;margin-left:3px;">[\1]</sup>', line_html)
        result.append(f'<p style="margin:0 0 18px;line-height:2.0;color:#2d2d2d;font-size:16px;">{line_html}</p>')
    return '\n'.join(result)


def build_toc_from_html(html: str) -> str:
    """HTML에서 <h2 id="secN">텍스트</h2> 추출 → TOC HTML 반환. 항목 2개 미만이면 빈 문자열."""
    items = re.findall(r'<h2[^>]+id="(sec\d+)"[^>]*>(.*?)</h2>', html, re.IGNORECASE)
    if len(items) < 2:
        return ''
    toc_css = (
        '<style>'
        '.toc-box{background:#f8f9fa;border:1px solid #dee2e6;border-radius:8px;'
        'padding:20px 24px;margin:24px 0 32px;max-width:700px;}'
        '.toc-box .toc-title{font-size:1.05em;font-weight:700;color:#2c3e50;margin-bottom:12px;}'
        '.toc-box ol{margin:0;padding-left:20px;}'
        '.toc-box ol li{margin:6px 0;}'
        '.toc-box a{color:#2980b9;text-decoration:none;font-size:.95em;}'
        '.toc-box a:hover{text-decoration:underline;}'
        '</style>'
    )
    li_tags = ''.join(
        f'<li><a href="#{sid}">{re.sub(r"<[^>]+>", "", text).strip()}</a></li>'
        for sid, text in items
    )
    toc_html = (
        f'<div class="toc-box">'
        f'<div class="toc-title">📋 목차</div>'
        f'<ol>{li_tags}</ol>'
        f'</div>'
    )
    return toc_css + toc_html


def embed_images_base64(html: str, image_paths: list) -> str:
    """base64 이미지를 </h2>/</h3> 태그 뒤에 순서대로 삽입"""
    if not image_paths:
        return html
    positions = []
    for tag in ('</h2>', '</h3>'):
        start = 0
        while True:
            pos = html.find(tag, start)
            if pos == -1:
                break
            positions.append(pos + len(tag))
            start = pos + 1
    positions.sort()

    inserts = []
    for idx, img_path in enumerate(image_paths):
        pos = positions[idx] if idx < len(positions) else len(html)
        inserts.append((pos, img_path))

    for pos, img_path in sorted(inserts, key=lambda x: x[0], reverse=True):
        try:
            with open(img_path, 'rb') as f:
                data = base64.b64encode(f.read()).decode()
            ext = 'png' if img_path.lower().endswith('.png') else 'jpeg'
            img_tag = (
                f'\n<figure style="margin:20px 0;border-radius:12px;overflow:hidden;'
                f'box-shadow:0 4px 18px rgba(0,0,0,.13);">'
                f'<img src="data:image/{ext};base64,{data}" '
                f'style="width:100%;max-width:700px;height:auto;display:block;">'
                f'</figure>'
            )
            html = html[:pos] + img_tag + html[pos:]
        except Exception as e:
            print(f"  이미지 embed 실패: {e}")
    return html


# ── STEP 2-1. 태그 자동 생성
def generate_tags(title: str, body: str) -> list:
    """제목과 본문을 분석해 티스토리 태그 자동 생성 (최대 10개)"""
    tags = ["테니스", "테니스스트링"]

    keyword_tag_map = {
        "폴리": "폴리에스터스트링",
        "나일론": "나일론스트링",
        "거트": "내추럴거트",
        "텐션": "스트링텐션",
        "스트링머신": "스트링머신",
        "알루파워": "럭실론알루파워",
        "럭실론": "럭실론",
        "바볼랏": "바볼랏스트링",
        "솔린코": "솔린코스트링",
        "요넥스": "요넥스스트링",
        "헤드": "헤드스트링",
        "윌슨": "윌슨스트링",
        "시그넘": "시그넘프로",
        "테크니화이버": "테크니화이버",
        "프리스트레칭": "프리스트레칭",
        "텐션유지": "텐션유지력",
        "클레이": "클레이코트",
        "잔디": "잔디코트",
        "하드코트": "하드코트",
        "동호인": "테니스동호인",
        "라켓": "테니스라켓",
        "스핀": "스핀테니스",
        "서브": "테니스서브",
        "롤랑가로스": "롤랑가로스",
        "윔블던": "윔블던",
        "US오픈": "US오픈",
        "호주오픈": "호주오픈",
    }

    combined = title + " " + body

    player_tag_map = {
        "알카라스": "알카라스", "알카라즈": "알카라스",
        "시너": "야닉시너", "야닉시너": "야닉시너",
        "조코비치": "조코비치", "나달": "나달", "페더러": "페더러",
        "메드베데프": "메드베데프", "즈베레프": "즈베레프",
        "루블레프": "루블레프", "루드": "카스퍼루드",
        "치치파스": "치치파스", "프리츠": "테일러프리츠",
        "셸튼": "벤셸튼", "드레이퍼": "잭드레이퍼",
        "무세티": "로렌초무세티", "루네": "홀거루네",
        "권순우": "권순우", "정현": "정현",
        "이덕희": "이덕희", "남지성": "남지성",
        "구연우": "구연우", "플리스코바": "플리스코바",
        "사발렌카": "사발렌카", "스비아텍": "스비아텍",
        "고프": "코코고프", "리바키나": "리바키나",
    }
    tournament_tag_map = {
        "롤랑가로스": "롤랑가로스", "윔블던": "윔블던",
        "US오픈": "US오픈", "호주오픈": "호주오픈",
        "마드리드오픈": "마드리드오픈", "바르셀로나오픈": "바르셀로나오픈",
        "이탈리아오픈": "이탈리아오픈", "몬테카를로": "몬테카를로",
        "인디안웰스": "인디안웰스", "마이애미오픈": "마이애미오픈",
        "신시내티": "신시내티오픈", "부산오픈": "부산오픈",
        "챌린저": "ATP챌린저",
    }
    for keyword, tag in {**player_tag_map, **tournament_tag_map}.items():
        if keyword in combined and tag not in tags:
            tags.append(tag)
        if len(tags) >= 10:
            break

    for keyword, tag in keyword_tag_map.items():
        if keyword in combined and tag not in tags:
            tags.append(tag)
        if len(tags) >= 10:
            break

    defaults = ["테니스용품", "테니스장비", "테니스정보", "테니스팁", "테니스블로그"]
    for t in defaults:
        if len(tags) >= 10:
            break
        if t not in tags:
            tags.append(t)

    print(f"🏷️  태그 {len(tags)}개: {', '.join(tags)}")
    return tags


# ── STEP 2-2. 카테고리 자동 감지
def detect_category(title: str, body: str) -> str:
    """제목과 본문 키워드로 카테고리 자동 감지"""
    combined = title + " " + body[:300]
    for keyword, category in CATEGORY_MAP.items():
        if keyword in combined:
            print(f"🗂️  카테고리 감지: '{keyword}' → {category}")
            return category
    default = list(CATEGORY_MAP.values())[0]
    print(f"🗂️  카테고리 기본값 사용: {default}")
    return default


# ── STEP 3. 이미지 HTML 빌더
def build_image_html(cdn_url: str, caption: str = "", style: str = "featured") -> str:
    """base64 URL로 이미지 HTML 블록 생성"""
    if not cdn_url:
        return ""
    if style == "featured":
        return f"""
<div style="margin:24px 0; border-radius:12px; overflow:hidden; box-shadow:0 4px 16px rgba(0,0,0,0.12);">
  <img src="{cdn_url}" alt="{caption}"
       style="width:100%; display:block; height:auto;">
</div>"""
    else:
        return f"""
<div style="margin:28px 0; border-radius:10px; overflow:hidden;
            box-shadow:0 2px 10px rgba(0,0,0,0.08); border:1px solid #e0e0e0;">
  <img src="{cdn_url}" alt="{caption}"
       style="width:100%; display:block; height:auto;">
</div>"""


def insert_images_into_body(body_html: str, cdn_url_list: list, min_gap: int = 600) -> str:
    """<h2>/<h3> 소제목마다 이미지 1장씩 삽입."""
    positions = sorted(
        [m.start() for m in re.finditer(r"<h2", body_html, re.IGNORECASE)] +
        [m.start() for m in re.finditer(r"<h3", body_html, re.IGNORECASE)]
    )
    candidates = positions[1:]

    insert_pairs = []
    last_pos = -min_gap
    img_idx = 0
    for pos in candidates:
        if img_idx >= len(cdn_url_list):
            break
        if pos - last_pos < min_gap:
            continue
        url = cdn_url_list[img_idx]
        if url:
            insert_pairs.append((pos, url))
            last_pos = pos
            img_idx += 1

    for pos, url in reversed(insert_pairs):
        html_img = build_image_html(url, caption="테니스 정보", style="inline")
        body_html = body_html[:pos] + html_img + body_html[pos:]

    return body_html


# ── STEP 4. 이미지 + 본문 조합
def wrap_in_theme(title: str, body_html: str, cdn_urls: list) -> str:
    """base64 URL 이미지를 본문에 삽입하고 테마 래핑"""
    img_featured = cdn_urls[0] if len(cdn_urls) > 0 else None
    img_mids     = [u for u in cdn_urls[1:] if u]

    html_featured    = build_image_html(img_featured, caption=title, style="featured") if img_featured else ""
    body_with_images = insert_images_into_body(body_html, img_mids)

    closing_img_path = _pick_female_closing_image()
    if closing_img_path:
        body_with_images += _img_to_html(closing_img_path)
        print(f'   🖼️  여성 선수 마감 이미지: {os.path.basename(closing_img_path)}')
        _save_used_tistory([os.path.basename(closing_img_path)])

    today = datetime.now().strftime("%Y년 %m월 %d일")

    header = f"""
<div style="background:linear-gradient(135deg,{_C_DARK} 0%,{_C_ACCENT} 100%);
            border-radius:16px;padding:34px 28px;margin-bottom:24px;text-align:center;
            box-shadow:0 6px 24px rgba(0,0,0,0.18);">
  <p style="color:#a5d6a7;font-size:12px;margin:0 0 8px;letter-spacing:3px;text-transform:uppercase;">🌍 글로벌 테니스 인텔리전스</p>
  <h1 style="color:#fff;font-size:24px;margin:0 0 10px;line-height:1.4;font-weight:800;">지금 테니스 세계에서 무슨 일이 일어나고 있을까요?</h1>
  <p style="color:#c8e6c9;font-size:13px;margin:0 0 6px;">전세계 미디어가 주목하는 테니스 소식을 한곳에 모았습니다.</p>
  <p style="color:#81c784;font-size:12px;margin:0;">📅 {today} 업데이트</p>
</div>"""

    return f"{header}\n{html_featured}\n<div style='padding:10px;'>\n{body_with_images}\n</div>\n{COMMON_FOOTER_HTML}"


# ── STEP 5. 로컬 이미지 → base64 데이터 URL 변환
def image_to_base64_url(image_path: str, max_width: int = 700) -> str:
    """로컬 이미지 파일을 base64 data URL로 변환 (항상 JPEG 700px/quality 65로 압축)"""
    mime = 'image/jpeg'
    try:
        from PIL import Image
        import io as _io
        img = Image.open(image_path).convert("RGB")
        if img.width > max_width:
            ratio = max_width / img.width
            img = img.resize((max_width, int(img.height * ratio)), Image.LANCZOS)
        buf = _io.BytesIO()
        img.save(buf, format='JPEG', quality=65)
        data = base64.b64encode(buf.getvalue()).decode('utf-8')
    except ImportError:
        # Pillow 없을 때: 원본 그대로 (최후 수단)
        with open(image_path, 'rb') as f:
            data = base64.b64encode(f.read()).decode('utf-8')
    print(f"  ✅ base64 변환: {os.path.basename(image_path)} ({len(data)//1024}KB)")
    return f"data:{mime};base64,{data}"


# ── STEP 6. 티스토리 포스팅
def post_to_tistory(title: str, body_html: str, image_paths: list, tags: list = None, category: str = '', thumbnail_path: str = None) -> bool:
    """티스토리에 HTML 본문을 TinyMCE로 주입하여 포스팅"""
    tags = tags or []
    title_for_input = re.sub(r'[^\w\s가-힣a-zA-Z0-9.,!\-()|&·~:?]', '', title).strip()
    if not title_for_input:
        title_for_input = title.strip()
    print('\n🔐 티스토리 카카오 로그인 중...')
    driver = get_driver()
    wait = WebDriverWait(driver, 35)
    try:
        _kakao_login(driver)

        driver.get(f'https://{TISTORY_BLOG}.tistory.com/manage/newpost/')
        time.sleep(6)
        print('   📄 에디터 로드 완료')

        try:
            alert = WebDriverWait(driver, 5).until(EC.alert_is_present())
            alert.dismiss()
            time.sleep(2)
        except Exception:
            pass

        # 제목 입력
        title_input = wait.until(EC.element_to_be_clickable((By.ID, 'post-title-inp')))
        driver.execute_script('arguments[0].focus();', title_input)
        title_input.send_keys(title_for_input)
        time.sleep(1)
        print(f'   📝 제목 입력: {title_for_input[:50]}')

        # 이미지 base64 변환
        print('   🔄 이미지 base64 변환 중...')
        cdn_urls = [image_to_base64_url(p) for p in image_paths]
        print(f'   ✅ 변환 완료: {len(cdn_urls)}장')

        # 카테고리 설정
        try:
            cat_btn = driver.execute_script("""
                var btns = document.querySelectorAll('button');
                for (var b of btns) {
                    if (b.textContent.trim().includes('카테고리')) return b;
                }
                var spans = document.querySelectorAll('i.mce-txt, span.mce-txt');
                for (var s of spans) {
                    if (s.textContent.trim().includes('카테고리'))
                        return s.closest('button') || s.parentElement;
                }
                return null;
            """)
            if cat_btn:
                driver.execute_script('arguments[0].click();', cat_btn)
                time.sleep(2)
                cat_item = driver.execute_script(f"""
                    var all = document.querySelectorAll('ul li, div[role="option"], div[role="listbox"] div, .tit_category, .list_category li');
                    for (var el of all) {{
                        if (el.textContent.trim() === '{category}') return el;
                    }}
                    for (var el of all) {{
                        if (el.textContent.trim().includes('{category}')) return el;
                    }}
                    return null;
                """)
                if cat_item:
                    driver.execute_script('arguments[0].click();', cat_item)
                    print(f'   🗂️  카테고리 설정: {category}')
                else:
                    print(f'   ⚠️ 카테고리 항목 못 찾음: {category}')
                    driver.save_screenshot(str(_BASE / 'category_debug.png'))
            else:
                print('   ⚠️ 카테고리 버튼 없음')
            time.sleep(1)
        except Exception as e:
            print(f'   ⚠️ 카테고리 설정 실패: {e}')

        # 최종 HTML 빌드 후 주입
        final_html = wrap_in_theme(title, body_html, cdn_urls)
        # HTML 크기 상한: 4MB 초과 시 이미지를 뒤에서부터 제거하며 재빌드
        _MAX_HTML_BYTES = 4 * 1024 * 1024
        while len(final_html.encode('utf-8')) > _MAX_HTML_BYTES and cdn_urls:
            cdn_urls.pop()
            print(f'   ⚠️ HTML {len(final_html)//1024}KB 초과 → 이미지 {len(cdn_urls)}장으로 줄임')
            final_html = wrap_in_theme(title, body_html, cdn_urls)
        print(f'   💉 HTML 본문 주입 중... ({len(final_html) // 1024} KB)')

        inserted = False
        has_script = '<script' in final_html.lower()

        if not has_script:
            try:
                driver.execute_script("""
                    var ed = tinymce.activeEditor
                           || tinymce.get('editor-tistory')
                           || (tinymce.editors && tinymce.editors[0]);
                    ed.setContent(arguments[0], {format: 'raw'});
                    ed.selection.select(ed.getBody(), true);
                    ed.selection.collapse(false);
                    ed.fire('change');
                    ed.focus();
                """, final_html)
                inserted = True
                print('   ✅ TinyMCE setContent 성공')
            except Exception as e:
                print(f'   ⚠️ TinyMCE API 실패: {e}')

        if not inserted:
            try:
                iframe = wait.until(EC.presence_of_element_located((By.ID, 'editor-tistory_ifr')))
                driver.switch_to.frame(iframe)
                body_el = wait.until(EC.presence_of_element_located((By.ID, 'tinymce')))
                driver.execute_script('arguments[0].innerHTML = arguments[1];', body_el, final_html)
                driver.execute_script("""
                    var b = document.getElementById('tinymce');
                    var r = document.createRange();
                    r.selectNodeContents(b);
                    r.collapse(false);
                    var s = window.getSelection();
                    s.removeAllRanges(); s.addRange(r); b.focus();
                """)
                driver.switch_to.default_content()
                driver.execute_script("""
                    var ed = tinymce.activeEditor || (tinymce.editors && tinymce.editors[0]);
                    if (ed) { ed.fire('change'); ed.focus(); }
                """)
                inserted = True
                print('   ✅ iframe innerHTML 주입 성공')
            except Exception as e:
                driver.switch_to.default_content()
                print(f'   ⚠️ iframe 주입 실패: {e}')

        if not inserted:
            driver.save_screenshot(str(_BASE / 'insert_debug.png'))
            raise Exception('HTML 본문 주입 실패')

        time.sleep(6)

        # 태그 입력
        try:
            for tag in tags:
                for attempt in range(5):
                    try:
                        tag_input = WebDriverWait(driver, 10).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, 'input#tagText'))
                        )
                        driver.execute_script('arguments[0].scrollIntoView(true);', tag_input)
                        driver.execute_script('arguments[0].focus();', tag_input)
                        time.sleep(0.3)
                        tag_input.send_keys(tag)
                        time.sleep(0.3)
                        tag_input.send_keys(Keys.RETURN)
                        time.sleep(0.5)
                        break
                    except (StaleElementReferenceException, Exception):
                        time.sleep(1)
            print(f'   🏷️  태그 {len(tags)}개 입력 완료')
            time.sleep(1)
        except Exception as e:
            print(f'   ⚠️ 태그 입력 실패: {e}')

        # 발행 레이어 열기 — 창 유효성 먼저 확인
        time.sleep(2)
        try:
            _ = driver.current_url  # 창이 살아있는지 확인
        except Exception:
            # 창이 닫혔으면 남은 핸들로 전환 시도
            handles = driver.window_handles
            if handles:
                driver.switch_to.window(handles[-1])
                time.sleep(1)
            else:
                raise Exception('모든 브라우저 창이 닫혔습니다.')

        # publish-layer-btn 클릭 (CSS / JS / XPATH 3단계 폴백)
        _layer_opened = False
        for _sel in ['#publish-layer-btn', '[id*="publish"][id*="layer"]',
                     'button[data-role="publish"]', '.btn-publish-layer']:
            try:
                _btn = WebDriverWait(driver, 8).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, _sel))
                )
                driver.execute_script('arguments[0].click();', _btn)
                _layer_opened = True
                break
            except Exception:
                continue
        if not _layer_opened:
            # JS 텍스트 기반 폴백 — "완료" 포함
            _layer_opened = driver.execute_script("""
                var btns = document.querySelectorAll('button, a, div[role="button"]');
                var kws = ['완료', '발행', '게시', 'publish', '저장', '올리기'];
                for (var kw of kws) {
                    for (var b of btns) {
                        var txt = b.textContent.trim();
                        if (txt === kw || txt.endsWith(kw)) { b.click(); return true; }
                    }
                }
                return false;
            """) or False
        time.sleep(3)
        print('   📤 발행 레이어 열림')

        # 비공개 설정 — CSS / JS 텍스트 기반 3단계 폴백
        _priv_done = False
        for _attempt in range(2):
            # 1단계: CSS 셀렉터
            for priv_sel in [
                'input#visibility1', "input[value='0']", "input[value='private']",
                "input[name='visibility'][value='0']", "input[name='visibility'][value='private']",
                "label[for='visibility1']", "label[for='visibility-private']",
                "label[for='visibility_private']", "label[for='visibilityPrivate']",
            ]:
                try:
                    priv_btn = WebDriverWait(driver, 4).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, priv_sel))
                    )
                    driver.execute_script('arguments[0].click();', priv_btn)
                    print('   🔒 비공개 설정 완료')
                    _priv_done = True
                    break
                except Exception:
                    continue
            if _priv_done:
                break
            # 2단계: JS 텍스트 기반
            _priv_done = driver.execute_script("""
                var els = document.querySelectorAll('label, button, span, li, div[role="radio"]');
                for (var el of els) {
                    var txt = el.textContent.trim();
                    if (txt === '비공개' || txt === '비공개 저장' || txt === 'Private') {
                        el.click(); return true;
                    }
                }
                return false;
            """) or False
            if _priv_done:
                print('   🔒 비공개 설정 완료 (JS 텍스트)')
                break
            time.sleep(2)
        time.sleep(1)

        # 대표 이미지 설정
        thumb_src = thumbnail_path or (image_paths[0] if image_paths else None)
        if thumb_src and Path(thumb_src).exists():
            try:
                thumb_btn = driver.execute_script("""
                    var btns = document.querySelectorAll(
                        'button, label, a, div[role="button"], span[role="button"]');
                    for (var b of btns) {
                        var txt = b.textContent.trim();
                        if (txt.includes('대표') || txt.includes('썸네일') ||
                            txt.includes('커버') || txt.includes('선택 하기') ||
                            txt.includes('선택하기') || txt.includes('이미지 등록') ||
                            txt.includes('이미지등록') || txt.includes('사진 추가') ||
                            txt.includes('사진추가')) return b;
                    }
                    return null;
                """)
                if not thumb_btn:
                    thumb_btn = driver.execute_script("""
                        return document.querySelector(
                            'button.btn-thumbnail, button.thumbnail-btn, ' +
                            'label.thumbnail-upload, #publish-thumbnail, ' +
                            '.publish-thumbnail button, .layer-publish-thumbnail button, ' +
                            '.cover-img button, .thumb-area button, ' +
                            'button[data-type="thumbnail"], label[for*="thumb"], label[for*="cover"]'
                        );
                    """)
                if thumb_btn:
                    driver.execute_script('arguments[0].click();', thumb_btn)
                    time.sleep(1.5)

                file_input = driver.execute_script("""
                    var inputs = document.querySelectorAll('input[type="file"]');
                    for (var inp of inputs) {
                        var accept = inp.accept || '';
                        if (accept.includes('image') || accept === '') return inp;
                    }
                    return null;
                """)
                if file_input:
                    driver.execute_script(
                        "arguments[0].style.display='block'; arguments[0].style.visibility='visible';",
                        file_input
                    )
                    file_input.send_keys(str(Path(thumb_src).resolve()))
                    time.sleep(5)
                    print(f'   🖼️  대표 이미지 업로드: {os.path.basename(thumb_src)}')
                    try:
                        confirm_btn = driver.execute_script("""
                            var btns = document.querySelectorAll('button');
                            for (var b of btns) {
                                var t = b.textContent.trim();
                                if (t === '확인' || t === '적용' || t === '완료' || t === '선택')
                                    return b;
                            }
                            return null;
                        """)
                        if confirm_btn:
                            driver.execute_script('arguments[0].click();', confirm_btn)
                            time.sleep(1.5)
                    except Exception:
                        pass
                else:
                    print('   ⚠️ 대표 이미지 file input 없음')
            except Exception as e:
                print(f'   ⚠️ 대표 이미지 설정 실패: {e}')

        # 발행/저장 버튼 클릭
        save_clicked = False
        try:
            publish_btn = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.ID, 'publish-btn'))
            )
            driver.execute_script('arguments[0].click();', publish_btn)
            save_clicked = True
            print('   🚀 발행 완료! (#publish-btn)')
            try:
                alert = WebDriverWait(driver, 4).until(EC.alert_is_present())
                alert.accept()
            except Exception:
                pass
        except Exception as e:
            print(f'   ⚠️ #publish-btn 실패: {e}')

        if not save_clicked:
            try:
                publish_btn = WebDriverWait(driver, 5).until(
                    EC.element_to_be_clickable((
                        By.XPATH,
                        "//button[normalize-space(text())='비공개 저장' or "
                        "normalize-space(text())='공개 저장' or "
                        "normalize-space(text())='저장' or "
                        "normalize-space(text())='발행' or "
                        "normalize-space(text())='발행하기' or "
                        "normalize-space(text())='완료' or "
                        "normalize-space(text())='올리기']"
                    ))
                )
                publish_btn.click()
                save_clicked = True
                print('   🚀 발행 완료! (XPATH)')
                try:
                    alert = WebDriverWait(driver, 4).until(EC.alert_is_present())
                    alert.accept()
                except Exception:
                    pass
            except Exception as e:
                print(f'   ⚠️ XPATH 발행 실패: {e}')

        # 비공개 저장 버튼 JS 텍스트 직접 탐색 (추가 폴백)
        if not save_clicked:
            try:
                result = driver.execute_script("""
                    var kws = ['비공개 저장', '비공개저장', '저장', '발행', '완료', '올리기',
                               'save', 'publish', 'submit'];
                    var btns = document.querySelectorAll(
                        'button, input[type="button"], input[type="submit"], a[role="button"]');
                    for (var kw of kws) {
                        for (var b of btns) {
                            var txt = (b.textContent || b.value || '').trim();
                            if (txt === kw || txt.includes(kw)) {
                                b.click(); return txt;
                            }
                        }
                    }
                    return null;
                """)
                if result:
                    save_clicked = True
                    print(f'   🚀 발행 완료! (JS 텍스트: {result})')
                    try:
                        alert = WebDriverWait(driver, 4).until(EC.alert_is_present())
                        alert.accept()
                    except Exception:
                        pass
            except Exception as e:
                print(f'   ⚠️ JS 텍스트 발행 실패: {e}')

        if not save_clicked:
            # JS 폴백: 발행 레이어 내 버튼 직접 탐색
            try:
                result = driver.execute_script("""
                    var layer = document.querySelector(
                        '#publish-layer, .publish-layer, [class*="publish"][class*="layer"], ' +
                        '[id*="publish"][id*="layer"], .layer-publish, #layer-publish, ' +
                        '.tt_layer_wrap, .layer_wrap, [class*="layer_wrap"]'
                    );
                    var searchRoot = layer || document;
                    var btns = searchRoot.querySelectorAll('button, input[type="button"], input[type="submit"]');
                    var keywords = ['완료', '발행', '저장', '공개', 'publish', 'save', 'submit', '올리기'];
                    for (var b of btns) {
                        var txt = (b.textContent || b.value || '').trim();
                        for (var kw of keywords) {
                            if (txt.includes(kw)) { b.click(); return txt; }
                        }
                    }
                    if (btns.length > 0) {
                        var last = btns[btns.length - 1];
                        last.click();
                        return 'last:' + (last.textContent || last.value || '').trim();
                    }
                    return null;
                """)
                if result:
                    save_clicked = True
                    print(f'   🚀 발행 완료! (JS 폴백: {result})')
                    try:
                        alert = WebDriverWait(driver, 4).until(EC.alert_is_present())
                        alert.accept()
                    except Exception:
                        pass
            except Exception as e:
                print(f'   ⚠️ JS 폴백 실패: {e}')

        if not save_clicked:
            print('   ❌ 발행 버튼 클릭 실패 — 포스팅 중단')
            return False

        def _dismiss_any_alert(d):
            """열린 alert을 닫고 텍스트 반환. alert 없으면 None."""
            try:
                al = WebDriverWait(d, 3).until(EC.alert_is_present())
                txt = al.text
                al.accept()
                return txt
            except Exception:
                return None

        def _is_success_url(d):
            # alert이 먼저 떠 있으면 닫은 뒤 URL 판단
            try:
                al = d.switch_to.alert
                txt = al.text
                al.accept()
                if '실패' in txt or 'fail' in txt.lower() or '오류' in txt:
                    print(f'   ⚠️ 발행 alert: {txt}')
                    return False
            except Exception:
                pass
            try:
                url = d.current_url
                return (
                    'manage/posts' in url
                    or ('manage/newpost' in url and '?' in url)
                    or (TISTORY_BLOG + '.tistory.com/' in url and 'manage' not in url)
                )
            except Exception:
                return False

        # 발행 직후 alert 먼저 처리
        _alert_txt = _dismiss_any_alert(driver)
        if _alert_txt:
            if '실패' in _alert_txt or '오류' in _alert_txt:
                print(f'   ❌ Tistory 발행 오류 alert: {_alert_txt}')
                print('   🔄 5초 후 재시도...')
                time.sleep(5)
                try:
                    driver.execute_script(
                        "var b = document.getElementById('publish-btn');"
                        "if(b) b.click();"
                    )
                    time.sleep(3)
                    _dismiss_any_alert(driver)
                except Exception:
                    pass
            else:
                print(f'   ℹ️ alert 처리: {_alert_txt}')

        try:
            WebDriverWait(driver, 60).until(_is_success_url)
            print(f'   ✅ 포스팅 성공! {driver.current_url}')
        except Exception:
            time.sleep(10)
            _dismiss_any_alert(driver)
            if _is_success_url(driver):
                print(f'   ✅ 포스팅 완료 (URL: {driver.current_url})')
            else:
                # 재시도: 발행 버튼 한 번 더 클릭
                print('   🔄 발행 재시도 중...')
                try:
                    btn = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.ID, 'publish-btn'))
                    )
                    btn.click()
                except Exception:
                    try:
                        driver.execute_script(
                            "document.getElementById('publish-btn').click()"
                        )
                    except Exception:
                        pass
                time.sleep(15)
                _dismiss_any_alert(driver)
                if _is_success_url(driver):
                    print(f'   ✅ 재시도 성공! {driver.current_url}')
                else:
                    try:
                        print(f'   ❌ 포스팅 실패 — URL: {driver.current_url}')
                    except Exception:
                        print('   ❌ 포스팅 실패 — URL 확인 불가')
                    return False

        return True

    except Exception as e:
        print(f'\n❌ 오류: {e}')
        import traceback
        traceback.print_exc()
        try:
            driver.save_screenshot(str(_BASE / 'error_debug.png'))
        except Exception:
            pass
        return False
    finally:
        time.sleep(5)
        driver.quit()


def load_from_txt(txt_path: str) -> tuple:
    """txt 파일에서 제목·본문·태그 추출."""
    with open(txt_path, "r", encoding="utf-8") as f:
        lines = f.readlines()

    title = ""
    content_lines = []
    tag_list = []
    title_found = False

    for line in lines:
        stripped = line.rstrip("\n").strip()

        if not title_found:
            if stripped:
                title = stripped
                title_found = True
            continue

        if stripped.startswith("#") and all(tok.startswith("#") for tok in stripped.split()):
            tag_list = [tok.lstrip("#") for tok in stripped.split()]
            continue

        if stripped == "---":
            content_lines.append("")
            continue

        if stripped.startswith("## "):
            content_lines.append(stripped[3:])
            continue

        content_lines.append(stripped)

    content = "\n".join(content_lines).strip()
    return title, content, tag_list


# ── 메인 실행
if __name__ == "__main__":
    print("=" * 60)
    print("🎾 테니스 자동 포스팅 (엑셀 본문 + 로컬 이미지 버전)")
    print("=" * 60)

    txt_arg = sys.argv[1] if len(sys.argv) > 1 and sys.argv[1].endswith(".txt") else None

    if txt_arg:
        if not os.path.exists(txt_arg):
            print(f"❌ txt 파일을 찾을 수 없습니다: {txt_arg}")
            sys.exit(1)
        with open(txt_arg, 'r', encoding='utf-8') as _f:
            raw_text = _f.read()

        # [POSTMETA] 파싱 및 제거
        _postmeta = _parse_postmeta(raw_text)
        _forced_players = _resolve_postmeta_players(_postmeta) if _postmeta else None
        if _forced_players:
            print(f"   🎯 POSTMETA 선수: {_forced_players}")
        raw_text = re.sub(r'\[POSTMETA\].*?\[/POSTMETA\]\s*\n?', '', raw_text, flags=re.DOTALL | re.IGNORECASE)

        # [SEO 메타정보] 블록 처리: 제목 추출 + 메타 블록 제거 후 본문만 남김
        _seo_title = ""
        _seo_tags_raw = ""
        _body_lines = []
        _in_meta = False
        for _li in raw_text.splitlines():
            _s = _li.strip()
            if re.match(r'^\[SEO\s*메타정보\]', _s, re.IGNORECASE):
                _in_meta = True
                continue
            if _in_meta:
                if _s.startswith('제목:'):
                    _seo_title = _s[3:].strip()
                elif _s.startswith('태그:'):
                    _seo_tags_raw = _s[3:].strip()
                elif not _s:
                    _in_meta = False
                continue
            _body_lines.append(_li)
        # 번호 섹션(1. 제목) → ## 제목 변환
        _body_lines = [
            re.sub(r'^(\d+)\.\s+(.+)$', r'## \2', _bl) for _bl in _body_lines
        ]
        raw_text_for_html = '\n'.join(_body_lines).strip()

        title, _, txt_tags = load_from_txt(txt_arg)
        # [SEO 메타정보]가 있으면 거기서 추출한 제목 우선
        if _seo_title:
            title = _seo_title
        if not title:
            print("❌ txt 파일에서 제목을 읽지 못했습니다.")
            sys.exit(1)
        # 태그 파싱 (#태그 형식)
        if _seo_tags_raw and not txt_tags:
            txt_tags = [t.lstrip('#').strip() for t in _seo_tags_raw.split() if t.startswith('#')]

        print(f"📄 txt 파일 사용: {os.path.basename(txt_arg)}")
        print(f"   제목: {title[:50]}")
        raw_body = md_to_html(raw_text_for_html)
        toc = build_toc_from_html(raw_body)
        if toc:
            raw_body = toc + raw_body
            print('   📋 목차(TOC) 삽입 완료')
        image_paths = select_images(title, raw_body, count=5, forced_players=_forced_players)
        thumbnail_path = image_paths[0] if image_paths else None

        tags = generate_tags(title, raw_body)
        if txt_tags:
            combined = txt_tags + [t for t in tags if t not in txt_tags]
            tags = combined[:10]
            print(f"   📌 txt 태그 우선 적용: {txt_tags}")
        category = detect_category(title, raw_text)

        ok = post_to_tistory(title, raw_body, image_paths=image_paths, tags=tags,
                             category=category, thumbnail_path=thumbnail_path)
        _save_used_tistory([os.path.basename(p) for p in image_paths])

    else:
        row_num, title, raw_body = read_post_from_excel()
        txt_tags = []

        if title and raw_body:
            image_paths = select_images(title, raw_body, count=5)
            tags = generate_tags(title, raw_body)
            category = detect_category(title, raw_body)

            ok = post_to_tistory(title, raw_body, image_paths, tags=tags, category=category)
            _save_used_tistory([os.path.basename(p) for p in image_paths])
            if row_num is not None:
                save_used_post(row_num)
        else:
            print("⚠️  데이터 읽기 실패로 중단합니다.")

    print("=" * 60)